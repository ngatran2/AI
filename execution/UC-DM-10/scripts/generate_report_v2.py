import os, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

src = r"d:\AI\Newtemplate\testcases\UC-DM-10\UC-DM-10_airline-category_testcases-hl_20260511_v1.xlsx"
out_dir = r"d:\AI\Newtemplate\execution\UC-DM-10\reports"
os.makedirs(out_dir, exist_ok=True)
out_file = os.path.join(out_dir, "res_UC-DM-10_airline-category_testcases-hl_res_20260511_v2.xlsx")

shutil.copy2(src, out_file)
wb = openpyxl.load_workbook(out_file)

results = {
    "TC_UC-DM-10_GUI_01": {
        "status": "PASS",
        "actual": "1. [UI] Bảng hiển thị 20/38 bản ghi, 2 cột dữ liệu: 'Mã hãng bay', 'Hãng bay' + cột 'Hành động'.\n2. [UI] Phân trang hiển thị 'Tổng: 38 - Bản ghi', nút 'Trang trước'/'Trang sau'.\n3. [UI] Nút '+ Thêm mới' hiển thị góc trên phải.\n4. [UI] Icon Sửa (✏️) và Xóa (🗑️) hiển thị đầy đủ mỗi dòng.",
        "note": "Environment: Staging (lams-fe.sotatek.works). Column headers khác TC: 'Mã hãng bay' (không phải 'Mã') và 'Hãng bay' (không phải 'Tên'). Cột label chính xác hơn."
    },
    "TC_UC-DM-10_FUNC_01": {
        "status": "PASS",
        "actual": "1. [UI] Popup 'Thêm hãng bay' mở khi nhấn '+ Thêm mới'. Fields: 'Mã hãng bay *', 'Hãng bay *'. Buttons: 'Đóng', 'Thêm mới'.\n2. [UI] Nhập Mã='T11', Tên='Auto T11 Test', nhấn 'Thêm mới'.\n3. [UI] Toast thành công xuất hiện. Popup đóng.\n4. [UI] Bản ghi T11 xuất hiện đầu danh sách. Tổng tăng từ 38 lên 39.",
        "note": "Data: Mã=T11, Tên=Auto T11 Test. Nút xác nhận là 'Thêm mới' (KHÔNG phải 'Lưu' như TC ghi). Toast text: 'Thêm mới T11 - Auto T11 Test thành công.'"
    },
    "TC_UC-DM-10_FUNC_02": {
        "status": "PASS",
        "actual": "1. [UI] Nhập Mã='T11' (trùng), Tên='Duplicate Test'.\n2. [UI] Nhấn 'Thêm mới'.\n3. [UI] Error Toast xuất hiện: 'Mã hãng bay đã tồn tại trong hệ thống.'.\n4. [UI] Popup vẫn mở, không lưu dữ liệu trùng.",
        "note": "Validation trùng mã hoạt động đúng. Error toast hiển thị chính xác như Expected Result."
    },
    "TC_UC-DM-10_FUNC_03": {
        "status": "PASS",
        "actual": "1. [UI] Nhấn icon Sửa (✏️) dòng T11 → Popup 'Chi tiết hãng bay' mở.\n2. [UI] Trường 'Mã hãng bay' bị disabled/read-only (nền xám, không nhập được). ✅\n3. [UI] Xóa tên cũ, nhập 'T11 Updated Name'. Nhấn 'Cập nhật'.\n4. [UI] Success Toast xuất hiện: 'Cập nhật T11 - T11 Updated Name thành công.'.\n5. [UI] Danh sách hiển thị tên mới.",
        "note": "Popup edit title='Chi tiết hãng bay'. Buttons: 'Xóa' và 'Cập nhật' (KHÔNG phải 'Lưu'). Mã field confirmed disabled."
    },
    "TC_UC-DM-10_FUNC_04": {
        "status": "PASS",
        "actual": "1. [UI] Nhấn icon Xóa (🗑️) dòng T11 → Popup 'Xác nhận' xuất hiện: 'Bạn có chắc chắn muốn xóa hãng bay T11 - T11 Updated Name khỏi danh mục?'.\n2. [UI] Nhấn 'Xác nhận'.\n3. [UI] Success Toast xuất hiện.\n4. [UI] T11 biến mất khỏi danh sách. Tổng giảm về 38.",
        "note": "Soft Delete confirmed. Confirmation dialog text khác requirement (hiển thị tên cụ thể thay vì generic)."
    },
    "TC_UC-DM-10_FUNC_05": {
        "status": "FAIL",
        "actual": "1. [UI] Nhấn icon Xóa dòng LHK → Popup xác nhận: 'Bạn có chắc chắn muốn xóa hãng bay LHK - Vietnam Airlines khỏi danh mục?'.\n2. [UI] Nhấn 'Xác nhận'.\n3. [UI] Hệ thống XÓA THÀNH CÔNG thay vì chặn. LHK biến mất khỏi danh sách.\n4. EXPECTED: Error Toast 'Không thể xóa khi hãng bay và loại thẻ liên quan đang được sử dụng.' → KHÔNG xuất hiện.",
        "note": "[RCA: R2 - Requirement Gap] LHK không được liên kết với Scanning Session/Cấu hình điều kiện trên staging. TC pre-condition yêu cầu data state có hãng bay đang được sử dụng, nhưng staging không có hãng nào bị ràng buộc. Cần chuẩn bị test data phù hợp hoặc xác nhận lại requirement."
    },
    "TC_UC-DM-10_FUNC_06": {
        "status": "PASS",
        "actual": "1. [UI] Thêm mới hãng SY1 - Sync Test Airline thành công ở module Hãng.\n2. [UI] Navigate sang Danh mục > Loại thẻ.\n3. [UI] Nhấn '+ Thêm mới' → Popup 'Thêm loại thẻ' mở.\n4. [UI] Có dropdown 'Hãng bay *' (placeholder: 'Chọn hãng bay').\n5. [UI] Dropdown chứa giá trị SY1 để chọn. ✅ Đồng bộ OK.",
        "note": "Integration Cross-module sync verified. Dropdown Hãng bay tại Loại thẻ hiển thị dữ liệu mới real-time."
    },
    "TC_UC-DM-10_FUNC_07": {
        "status": "PASS",
        "actual": "1. [UI] Đăng nhập với role View (nga.tran2+50@sotatek.com).\n2. [UI] Navigate Danh mục > Hãng.\n3. [UI] Nút '+ Thêm mới' bị ẨN hoàn toàn (không hiển thị).\n4. [UI] Cột 'Hành động' chỉ hiển thị icon Xem (👁️), KHÔNG có icon Sửa/Xóa.\n5. [UI] Người dùng chỉ xem danh sách và tìm kiếm. ✅",
        "note": "RBAC View-only verified. Account: nga.tran2+50@sotatek.com. Login via 2FA OTP thành công."
    }
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
black_font = Font(name="Segoe UI", size=11, color="000000")
bold_font = Font(name="Segoe UI", size=11, color="000000", bold=True)
wrap = Alignment(wrap_text=True, vertical="top")

for sheet_name in ["GUI", "FUNCTION"]:
    ws = wb[sheet_name]
    headers = {"H": "Status", "I": "Actual Result", "J": "Note"}
    for col, title in headers.items():
        cell = ws[f"{col}1"]
        if not cell.value:
            cell.value = title
            cell.font = bold_font

    for row_num in range(2, ws.max_row + 1):
        tc_id_cell = ws.cell(row=row_num, column=1)
        tc_id = str(tc_id_cell.value).strip() if tc_id_cell.value else ""
        if not tc_id:
            continue
        if tc_id in results:
            r = results[tc_id]
            status_cell = ws.cell(row=row_num, column=8)
            actual_cell = ws.cell(row=row_num, column=9)
            note_cell = ws.cell(row=row_num, column=10)

            status_cell.value = r["status"]
            actual_cell.value = r["actual"]
            note_cell.value = r["note"]

            for c in [status_cell, actual_cell, note_cell]:
                c.font = black_font
                c.alignment = wrap

            if r["status"] == "PASS":
                status_cell.fill = green_fill
                status_cell.font = Font(name="Segoe UI", size=11, color="006100", bold=True)
            elif r["status"] == "FAIL":
                status_cell.fill = red_fill
                status_cell.font = Font(name="Segoe UI", size=11, color="9C0006", bold=True)
            else:
                status_cell.fill = gray_fill

            ws.row_dimensions[row_num].hidden = False
            ws.row_dimensions[row_num].height = 100

wb.save(out_file)
print("Report saved to", out_file)
