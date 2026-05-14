import os, sys, shutil
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

src = r"d:\AI\Newtemplate\testcases\UC-DM-10\UC-DM-10_airline-category_testcases-hl_20260511_v1.xlsx"
out_dir = r"d:\AI\Newtemplate\execution\UC-DM-10\reports"
os.makedirs(out_dir, exist_ok=True)
out_file = os.path.join(out_dir, "res_UC-DM-10_airline-category_testcases-hl_res_20260511_v1.xlsx")

shutil.copy2(src, out_file)
wb = openpyxl.load_workbook(out_file)

# Execution results mapping
results = {
    "TC_UC-DM-10_GUI_01": {
        "status": "PASS",
        "actual": "1. [UI] Bảng hiển thị 20/25 bản ghi với 2 cột (Mã, Tên) + cột Thao tác.\n2. [UI] Phân trang hiển thị Tổng: 25, nút Previous/Next.\n3. [UI] Nút Thêm mới, icon Sửa/Xóa hiển thị đầy đủ.",
        "note": "Simulated Mock Screen. UI verified via browser."
    },
    "TC_UC-DM-10_FUNC_01": {
        "status": "PASS",
        "actual": "1. [UI] Popup Thêm mới đóng lại sau khi nhấn Lưu.\n2. [UI] Success Toast 'Thêm mới [TEST01 - Hang bay tu dong] thành công.' xuất hiện góc dưới trái.\n3. [UI] Bản ghi mới xuất hiện trong danh sách.",
        "note": "Data: Mã=TEST01, Tên=Hang bay tu dong. Mock DB verified."
    },
    "TC_UC-DM-10_FUNC_02": {
        "status": "PASS",
        "actual": "1. [UI] Hệ thống chặn lưu.\n2. [UI] Error Toast 'Mã hãng bay đã tồn tại trong hệ thống.' xuất hiện.",
        "note": "Data: Mã=VN (đã tồn tại). Validation logic hoạt động đúng."
    },
    "TC_UC-DM-10_FUNC_03": {
        "status": "PASS",
        "actual": "1. [UI] Popup Sửa mở, trường Mã bị read-only.\n2. [UI] Success Toast 'Cập nhật [QH - Bamboo Airways Updated] thành công.' xuất hiện.\n3. [UI] Danh sách cập nhật tên mới.",
        "note": "Mã QH readOnly confirmed. Tên đổi thành 'Bamboo Airways Updated'."
    },
    "TC_UC-DM-10_FUNC_04": {
        "status": "PASS",
        "actual": "1. [UI] Popup xác nhận 'Bạn có chắc chắn muốn xóa?' hiển thị.\n2. [UI] Success Toast 'Xóa [VJ - VietJet Air] và các loại thẻ liên quan thành công.' xuất hiện.\n3. [UI] Hãng VJ biến mất khỏi danh sách.",
        "note": "Soft Delete verified. VJ removed from UI list."
    },
    "TC_UC-DM-10_FUNC_05": {
        "status": "PASS",
        "actual": "1. [UI] Popup xác nhận hiển thị, nhấn Xác nhận.\n2. [UI] Error Toast 'Không thể xóa khi hãng bay và loại thẻ liên quan đang được sử dụng.' xuất hiện.\n3. [UI] Hãng VN vẫn giữ nguyên trong danh sách.",
        "note": "Business constraint (linked data) enforced correctly."
    },
    "TC_UC-DM-10_FUNC_06": {
        "status": "SKIPPED",
        "actual": "Không thể kiểm tra. Mock screen không bao gồm module Quản lý loại thẻ với Dropdown tham chiếu.",
        "note": "Integration test requires full system. Deferred to real environment."
    },
    "TC_UC-DM-10_FUNC_07": {
        "status": "PASS",
        "actual": "1. [UI] Sau khi chuyển sang Role View-only, nút 'Thêm mới' bị ẩn.\n2. [UI] Tất cả icon Sửa/Xóa trong bảng bị ẩn hoàn toàn.\n3. [UI] Chỉ có thể xem danh sách.",
        "note": "RBAC verified via switchRole('view'). Admin buttons hidden."
    }
}

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
black_font = Font(name="Segoe UI", size=11, color="000000")
wrap = Alignment(wrap_text=True, vertical="top")

for sheet_name in ["GUI", "FUNCTION"]:
    ws = wb[sheet_name]
    # Header Guard
    headers = {"H": "Status", "I": "Actual Result", "J": "Note"}
    for col, title in headers.items():
        cell = ws[f"{col}1"]
        if not cell.value:
            cell.value = title
            cell.font = Font(name="Segoe UI", size=11, bold=True, color="000000")

    for row_num in range(2, ws.max_row + 1):
        tc_id_cell = ws.cell(row=row_num, column=1)
        tc_id = str(tc_id_cell.value).strip() if tc_id_cell.value else ""
        if not tc_id:
            continue
        if tc_id in results:
            r = results[tc_id]
            status_cell = ws.cell(row=row_num, column=8)
            actual_cell = ws.cell(row=row_num, column=9)
            note_cell = ws.cell(row=row_num, column=10)

            status_cell.value = r["status"]
            actual_cell.value = r["actual"]
            note_cell.value = r["note"]

            for c in [status_cell, actual_cell, note_cell]:
                c.font = black_font
                c.alignment = wrap

            if r["status"] == "PASS":
                status_cell.fill = green_fill
            elif r["status"] == "FAIL":
                status_cell.fill = red_fill
            else:
                status_cell.fill = gray_fill

            ws.row_dimensions[row_num].hidden = False
            ws.row_dimensions[row_num].height = 80

wb.save(out_file)
print("Report saved to", out_file)
